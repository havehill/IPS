import requests
from bs4 import BeautifulSoup
import openpyxl

locals = [[11, 1125], [12, 1216],[13,1308],[14,1410],[15,1505],[16,1605],[17,1705],[18,1801],[21,2144],[22,2218],[23,2314],[24,2416],[25,2515],[26,2622],[27,2724],[28,2822],[29,2902]]

fpath = r'C:\Users\SM-PC\Desktop\IPS\code\IPS_격리병동수\전국.xlsx'
wb = openpyxl.Workbook()
ws = wb.active  # 현재 활성화된 시트 선택
ws.cell(row=1, column=1).value = 'name'
ws.cell(row=1, column=2).value = 'loc'

row = 2
column = 3

for local in locals:
    city = local[0]
    last_num = local[1]

    

    for gu in range(int(str(city)+'01'), int(last_num)+1):
        response = requests.get(f'https://portal.nemc.or.kr:444/medi_info/dashboards/dash_total_emer_org_popup_for_egen.do?juso=&lon=&lat=&con=on&emogloca={city}&emogdstr={gu}&asort=A&asort=C&asort=D&rltmCd=O003&rltmCd=O004&rltmCd=O048&rltmCd=O049&rltmCd=O046&rltmCd=O047&rltmCd=O052&rltmCd=O053&rltmCd=O050&rltmCd=O051&rltmCd=O018&rltmCd=O025&rltmC19Cd=O054&rltmC19Cd=O055&rltmC19Cd=O056&rltmC19Cd=O057&rltmC19Cd=O058&rltmC19Cd=O059&afterSearch=map&theme=BLACK&refreshTime=60&spreadAllMsg=allClose&searchYn=Y')
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        links = soup.find_all(class_="dash_box")
        # print(links)
        for link in links:
            column = 3
            name = link['data-emogdesc']
            print(name)
            ws.cell(row=row, column=1).value = name
            loc = link.find(class_='info').span.text
            print(loc)
            ws.cell(row=row, column=2).value = loc

            try:
            
                #link > div > div > table > tbody > tr  > td
                tds = link.find_all('td')
                for td in tds:
                    col = td.select_one('div:nth-child(1)').text
                    data = td.select_one('div:nth-child(2)').text
                    col = ''.join(col.split())
                    data = ''.join(data.split())
                    print(col, data)
                    print(column) 
                    ws.cell(row=1, column=column).value = col
                    ws.cell(row=row, column=column).value = data
                    column = column + 1
                row = row + 1
            except:
                print(name+ loc + 'error발생----------------------------------------------------------------------')
        
wb.save(fpath)
        
